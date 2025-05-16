"""报告生成模块"""

import os
import base64
import re
import matplotlib.pyplot as plt
from .constants import Colors, EMA_PERIODS
import pandas as pd
from openpyxl.utils import get_column_letter

def remove_ansi_colors(text):
    """移除文本中的 ANSI 颜色代码
    Args:
        text (str): 包含 ANSI 颜色代码的文本
    Returns:
        str: 移除了 ANSI 颜色代码的文本
    """
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    return ansi_escape.sub('', text)

def save_analysis_plot(data, symbol, output_dir):
    """保存分析图表，采用现代简约的设计风格
    Args:
        data (pd.DataFrame): 股票数据
        symbol (str): 股票代码
        output_dir (str): 输出目录
    """
    try:
        # 导入样式模块
        from .styles import get_color_scheme
        
        # 获取配色方案
        colors = get_color_scheme()
        
        # 设置Matplotlib样式
        plt.style.use('seaborn-v0_8-whitegrid') # 使用更现代的样式
        
        # 导入字体管理器，确保字体正确加载
        import matplotlib.font_manager as fm
        
        # Set Matplotlib default font
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'Helvetica'] # 添加备选字体
        plt.rcParams['font.family'] = 'sans-serif'
        print('Using DejaVu Sans font for plots.')
        
        # 统一字体和符号配置
        plt.rcParams.update({
            'axes.unicode_minus': False,  # 正确显示负号
            'font.size': 12,              # 统一字体大小
            'axes.titlesize': 18,         # 标题字体大小 (增大)
            'axes.labelsize': 14,         # 坐标轴标签大小 (增大)
            'legend.fontsize': 11,        # 图例字体大小
            'xtick.labelsize': 10,        # X轴刻度标签大小
            'ytick.labelsize': 10,        # Y轴刻度标签大小
        })
        
        plt.rcParams['axes.facecolor'] = colors.get('background_light', '#FFFFFF') # 更亮的背景
        plt.rcParams['figure.facecolor'] = colors.get('background_light', '#FFFFFF')
        plt.rcParams['axes.edgecolor'] = colors.get('border_light', '#CCCCCC') # 更浅的边框
        plt.rcParams['axes.labelcolor'] = colors.get('text_strong', '#333333') # 更强的文本颜色
        plt.rcParams['xtick.color'] = colors.get('text_medium', '#555555') # 中等文本颜色
        plt.rcParams['ytick.color'] = colors.get('text_medium', '#555555')
        
        # 创建图表
        fig, ax1 = plt.subplots(figsize=(16, 8)) # 略微调整尺寸
        
        # 创建副图 (RSI)
        ax2 = ax1.twinx()
        
        # 格式化日期
        data_plot = data.copy()
        data_plot.index = pd.to_datetime(data_plot.index).strftime('%Y-%m-%d')
        
        # 绘制价格
        ax1.plot(data_plot.index, data_plot['Close'], label='Price', 
                color=colors.get('primary', '#007ACC'), 
                linewidth=2.5, 
                alpha=0.9)
        
        # 定义要绘制的EMA周期和对应的颜色、样式
        # 用户要求: Price/EMA5, EMA10, EMA20, 和 EMA200
        emas_to_plot = {
            'EMA_5': {'label': 'EMA 5', 'color': colors.get('chart_green', '#2ECC71'), 'style': '-', 'lw': 1.8},
            'EMA_10': {'label': 'EMA 10', 'color': colors.get('chart_blue', '#3498DB'), 'style': '-', 'lw': 1.8},
            'EMA_20': {'label': 'EMA 20', 'color': colors.get('chart_purple', '#9B59B6'), 'style': '-', 'lw': 1.8},
            'EMA_200': {'label': 'EMA 200', 'color': colors.get('chart_red', '#E74C3C'), 'style': '--', 'lw': 2.0} # EMA200用虚线突出
        }
        
        chart_colors = colors.get('chart_colors', ['#2ECC71', '#3498DB', '#9B59B6', '#F1C40F', '#E67E22', '#E74C3C'])

        # 绘制选定的EMA线
        for i, (ema_col, props) in enumerate(emas_to_plot.items()):
            if ema_col in data_plot.columns:
                ax1.plot(data_plot.index, data_plot[ema_col], 
                        label=props['label'],
                        color=props['color'],
                        linestyle=props['style'],
                        linewidth=props['lw'],
                        alpha=0.85)
            else:
                print(f"Warning: {ema_col} not found in data for {symbol}")
        
        # 绘制RSI
        ax2.plot(data_plot.index, data_plot['RSI'], 
                label='RSI', 
                color=colors.get('info_dark', '#F39C12'), # 更鲜明的RSI颜色
                linewidth=1.8, 
                alpha=0.8)
        
        # RSI超买超卖线
        ax2.axhline(y=70, color=colors.get('negative_light', '#E74C3C'), linestyle=':', alpha=0.6, linewidth=1.5) # 虚点线
        ax2.axhline(y=30, color=colors.get('positive_light', '#2ECC71'), linestyle=':', alpha=0.6, linewidth=1.5) # 虚点线
        ax2.fill_between(data_plot.index, 70, 100, color=colors.get('negative_light', '#E74C3C'), alpha=0.1) # 使用基础颜色和alpha控制透明度
        ax2.fill_between(data_plot.index, 0, 30, color=colors.get('positive_light', '#2ECC71'), alpha=0.1) # 使用基础颜色和alpha控制透明度
        
        # 设置标题和自定义图表
        latest_date = pd.to_datetime(data_plot.index[-1]).strftime("%Y-%m-%d")
        ax1.set_title(f'{symbol} Technical Analysis - {latest_date}', 
                     fontsize=18, 
                     color=colors.get('text_strong', '#333333'),
                     fontweight='bold',
                     pad=25) # 增加标题和图表的间距
        
        # 网格设置 (更细的网格线)
        ax1.grid(True, alpha=0.3, color=colors.get('border_light', '#DDDDDD'), linestyle='--')
        ax2.grid(False)  # 关闭RSI区域的网格，减少视觉干扰
        
        # 格式化x轴
        # 自动选择最佳的刻度数量和位置
        ax1.xaxis.set_major_locator(plt.MaxNLocator(nbins=10, prune='both')) 
        plt.setp(ax1.xaxis.get_majorticklabels(), 
                rotation=30, # 调整旋转角度
                ha='right', 
                fontsize=10, 
                color=colors.get('text_medium', '#555555'))
        
        # 设置标签
        ax1.set_xlabel('Date', fontsize=14, color=colors.get('text_strong', '#333333'))
        ax1.set_ylabel('Price / EMA', fontsize=14, color=colors.get('text_strong', '#333333'))
        ax2.set_ylabel('RSI', fontsize=14, color=colors.get('info_dark', '#F39C12'))
        
        # 设置Y轴范围
        ax2.set_ylim(0, 100)
        
        # 移除顶部和右侧的 spines，使图表更简洁
        ax1.spines['top'].set_visible(False)
        ax1.spines['right'].set_visible(False)
        ax2.spines['top'].set_visible(False)
        # ax2.spines['right'].set_visible(False) # RSI Y轴保留

        # 调整左侧和底部spines的颜色
        ax1.spines['left'].set_color(colors.get('border_medium', '#AAAAAA'))
        ax1.spines['bottom'].set_color(colors.get('border_medium', '#AAAAAA'))
        ax2.spines['left'].set_color(colors.get('border_medium', '#AAAAAA')) # ax2的左spine其实是ax1的
        ax2.spines['right'].set_color(colors.get('border_medium', '#AAAAAA'))
        ax2.spines['bottom'].set_color(colors.get('border_medium', '#AAAAAA'))
        
        # 添加图例 (合并图例，放置在图表外部或最佳位置)
        lines, labels = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        # ax1.legend(lines + lines2, labels + labels2, loc='upper left', frameon=True, 
        #           facecolor=colors.get('background_semi_transparent', 'rgba(255, 255, 255, 0.8)'), 
        #           edgecolor=colors.get('border_light', '#DDDDDD'), framealpha=0.9, ncol=2) # 两列图例

        # 将图例放置在图表下方，水平排列
        fig.legend(lines + lines2, labels + labels2, loc='lower center', 
                   bbox_to_anchor=(0.5, -0.05), # 调整位置使其在图表下方
                   ncol=len(labels + labels2), # 根据图例数量自动调整列数
                   frameon=False, # 无边框
                   fontsize=11)

        # 添加水印
        plt.figtext(0.98, 0.01, 'Stock Analysis by TraeAI', 
                   fontsize=9, color=colors.get('text_light', '#999999'), ha='right') # 水印放置在右下角
        
        # 自动调整布局，为图例留出空间
        plt.tight_layout(rect=[0, 0.05, 1, 0.95]) # rect=[left, bottom, right, top]
        
        # 保存图表
        plot_filename = os.path.join(output_dir, f'{symbol}_analysis_plot.png')
        plt.savefig(plot_filename,
                   bbox_inches='tight', 
                   dpi=150, # 提高DPI以获得更清晰的图像
                   facecolor=fig.get_facecolor())
        plt.close(fig) # 关闭图表，释放内存
        print(f"Modern plot saved to {plot_filename}")
        
    except Exception as e:
        print(f"{Colors.RED}Error saving modern analysis plot for {symbol}: {str(e)}{Colors.END}")

def format_value(value):
    """格式化数值
    Args:
        value: 要格式化的值
    Returns:
        str: 格式化后的字符串
    """
    if isinstance(value, float):
        return f"{value:.2f}"
    if isinstance(value, pd.Series):
        return f"{value.iloc[0]:.2f}"
    return str(value)

def save_to_excel(results, output_dir):
    """保存数据到Excel文件
    Args:
        results (list): 分析结果列表
        output_dir (str): 输出目录
    """
    try:
        # 创建一个Excel写入器
        excel_file = os.path.join(output_dir, 'stock_analysis.xlsx')
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 创建概览表
            overview_data = []
            for result in results:
                if not result.get('error'):
                    overview_data.append({
                        '股票代码': result['symbol'],
                        '当前价格': format_value(result['price']),
                        '价格变化': format_value(result['price_change']),
                        '价格变化率(%)': format_value(result['price_change_pct']),
                        'RSI': format_value(result['rsi']),
                        '警报数量': len(result.get('alert_details', []))
                    })
            
            # 保存概览表
            if overview_data:
                overview_df = pd.DataFrame(overview_data)
                overview_df.set_index('股票代码', inplace=True)
                
                # 格式化数字列为两位小数
                float_columns = ['当前价格', '价格变化', '价格变化率(%)', 'RSI']
                for col in float_columns:
                    try:
                        overview_df[col] = pd.to_numeric(overview_df[col])
                    except Exception:
                        # 跳过无法转换的列
                        pass
                
                overview_df.to_excel(writer, sheet_name='概览', float_format='%.2f')
                
                # 获取工作表
                worksheet = writer.sheets['概览']
                
                # 定义样式
                from openpyxl.styles import PatternFill, Font
                red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
                
                # 为每个单元格添加条件格式化
                for row in range(2, worksheet.max_row + 1):  # 从第2行开始（跳过标题）
                    # 价格变化和变化率
                    for col in ['C', 'D']:  # B是股票代码，C是价格变化，D是变化率
                        cell = worksheet[f'{col}{row}']
                        if cell.value:
                            value = float(cell.value)
                            if value > 0:
                                cell.fill = green_fill
                            elif value < 0:
                                cell.fill = red_fill
                    
                    # RSI
                    rsi_cell = worksheet[f'E{row}']
                    if rsi_cell.value:
                        rsi = float(rsi_cell.value)
                        if rsi >= 70:
                            rsi_cell.fill = red_fill
                        elif rsi <= 30:
                            rsi_cell.fill = green_fill
                    
                    # 警报数量
                    alert_cell = worksheet[f'F{row}']
                    if alert_cell.value and int(alert_cell.value) > 0:
                        alert_cell.fill = yellow_fill
            else:
                # 如果没有概览数据，创建一个空的概览表
                pd.DataFrame().to_excel(writer, sheet_name='概览')
            
            # 为每个股票创建详细数据表
            for result in results:
                if not result.get('error') and 'data' in result:
                    symbol = result['symbol']
                    data = result['data'].copy()
                    
                    # 格式化日期索引
                    data.index = data.index.strftime('%Y-%m-%d')
                    
                    # 重命名列
                    rename_map = {
                        'Open': '开盘价',
                        'High': '最高价',
                        'Low': '最低价',
                        'Close': '收盘价',
                        'Adj Close': '调整后收盘价',
                        'Volume': '成交量',
                        'RSI': 'RSI'
                    }
                    for period in EMA_PERIODS:
                        rename_map[f'EMA_{period}'] = f'{period}日均线'
                    
                    data.rename(columns=rename_map, inplace=True)
                    
                    # 保存到Excel，所有数字列保留两位小数
                    data.to_excel(writer, sheet_name=symbol, float_format='%.2f')
                    
                    # 获取工作表
                    worksheet = writer.sheets[symbol]
                    
                    # 为每个单元格添加条件格式化
                    for row in range(2, worksheet.max_row + 1):
                        # RSI
                        rsi_col = None
                        for idx, col in enumerate(worksheet[1], 1):  # 第一行是标题
                            if col.value == 'RSI':
                                rsi_col = idx
                                break
                        
                        if rsi_col:
                            rsi_cell = worksheet.cell(row=row, column=rsi_col)
                            if rsi_cell.value:
                                rsi = float(rsi_cell.value)
                                if rsi >= 70:
                                    rsi_cell.fill = red_fill
                                elif rsi <= 30:
                                    rsi_cell.fill = green_fill
            
            # 设置列宽
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(worksheet.columns, 1):
                    worksheet.column_dimensions[get_column_letter(idx)].width = 15
                    
    except Exception as e:
        print(f"{Colors.RED}Error saving Excel file: {str(e)}{Colors.END}")

def generate_html_report(results, output_dir):
    """生成HTML报告
    Args:
        results (list): 分析结果列表
        output_dir (str): 输出目录
    """
    try:
        # 保存Excel文件
        save_to_excel(results, output_dir)
        
        # 导入样式模块
        from .styles import get_css_styles
        
        # 获取当前日期
        from datetime import datetime
        current_date = datetime.now().strftime("%Y年%m月%d日")
        
        # HTML头部
        html = f"""
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>股票分析报告</title>
            <link rel="preconnect" href="https://fonts.googleapis.com">
            <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
            <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
            <style>
                {get_css_styles()}
            </style>
        </head>
        <body>
            <h1>股票分析报告 <span style="font-weight: 400; font-size: 1.2rem; color: #8A919E; margin-left: 12px;">{current_date}</span></h1>
            
            <div class="card">
                <div class="card-header">
                    <h2 style="margin: 0;">市场概览</h2>
                </div>
                <div class="card-body">
                    <p>详细数据已保存到 <a href="stock_analysis.xlsx">stock_analysis.xlsx</a></p>
                    <table>
                        <thead>
                            <tr>
                                <th>股票代码</th>
                                <th>当前价格</th>
                                <th>价格变化</th>
                                <th>RSI</th>
                                <th>警报数量</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        # 添加概览表格数据
        for result in results:
            if result.get('error'):
                html += f"""
                <tr>
                    <td>{result['symbol']}</td>
                    <td colspan="4" style="color: #D82C0D;">Error: {result['error']}</td>
                </tr>
                """
                continue
                
            price_change = float(format_value(result['price_change']))
            price_change_class = 'up' if price_change > 0 else 'down'
            rsi_value = float(format_value(result['rsi']))
            rsi_class = ""
            if rsi_value >= 70:
                rsi_class = "down"  # 超买
            elif rsi_value <= 30:
                rsi_class = "up"    # 超卖
            
            alert_count = len(result.get('alert_details', []))
            
            html += f"""
            <tr>
                <td><strong>{result['symbol']}</strong></td>
                <td>{format_value(result['price'])}</td>
                <td class="{price_change_class}">{format_value(result['price_change'])} ({format_value(result['price_change_pct'])} %)</td>
                <td class="{rsi_class}">{format_value(result['rsi'])}</td>
                <td>{alert_count} {f'<span class="alert">{"警报" if alert_count > 0 else ""}</span>' if alert_count > 0 else ""}</td>
            </tr>
            """
        
        html += """
                        </tbody>
                    </table>
                </div>
            </div>
            
            <div class="card">
                <div class="card-header">
                    <h2 style="margin: 0;">详细分析</h2>
                </div>
            </div>
        """
        
        # 添加每个股票的详细分析
        for result in results:
            if result.get('error'):
                continue
                
            # 获取主要指标
            price = format_value(result['price'])
            price_change = format_value(result['price_change'])
            price_change_pct = format_value(result['price_change_pct'])
            rsi = format_value(result['rsi'])
            change_class = 'up' if float(price_change) > 0 else 'down'
            
            html += f"""
            <div class="card stock-detail">
                <div class="card-header">
                    <h3 style="margin: 0;">{result['symbol']} 详细分析</h3>
                </div>
                <div class="card-body">
                    <!-- 关键指标部分 -->
                    <div class="metrics-container">
                        <div class="metric-card">
                            <div class="metric-name">价格</div>
                            <div class="metric-value">{price}</div>
                        </div>
                        <div class="metric-card">
                            <div class="metric-name">变化</div>
                            <div class="metric-value {change_class}">{price_change}</div>
                            <div class="metric-change {change_class}">({price_change_pct} %)</div>
                        </div>
                        <div class="metric-card">
                            <div class="metric-name">RSI</div>
                            <div class="metric-value">{rsi}</div>
                            <div class="metric-change">{"超买" if float(rsi) >= 70 else "超卖" if float(rsi) <= 30 else "正常"}</div>
                        </div>
                        <div class="metric-card">
                            <div class="metric-name">警报</div>
                            <div class="metric-value">{len(result.get('alert_details', []))}</div>
                        </div>
                    </div>
                    
                    <!-- 分析图表 -->
                    <img src="{result['symbol']}_analysis_plot.png" alt="{result['symbol']} 分析图表">
            """
            
            # 添加警报信息
            if result.get('alert_details', []):
                html += """<div class="card" style="margin-top: 24px;">
                    <div class="card-header">
                        <h4 style="margin: 0;">警报信息</h4>
                    </div>
                    <div class="card-body">
                        <ul>"""
                for alert in result['alert_details']:
                    # 提取日期并检查是否在最近10天内
                    import re
                    from datetime import datetime, timedelta
                    date_match = re.search(r'发生于：(\d{4}-\d{2}-\d{2})', alert)
                    is_recent = False
                    if date_match:
                        alert_date = datetime.strptime(date_match.group(1), '%Y-%m-%d')
                        current_date = datetime.now()
                        days_diff = (current_date - alert_date).days
                        is_recent = days_diff <= 10
                    
                    # 清除ANSI颜色代码
                    clean_alert = remove_ansi_colors(alert)
                    
                    # 替换常见的中英文混排情况，添加空格
                    import re
                    # 在英文字母/数字和中文之间添加空格
                    clean_alert = re.sub(r'([a-zA-Z0-9])([\u4e00-\u9fff])', r'\1 \2', clean_alert)
                    clean_alert = re.sub(r'([\u4e00-\u9fff])([a-zA-Z0-9])', r'\1 \2', clean_alert)
                    # 确保数字与百分号之间有空格
                    clean_alert = re.sub(r'([0-9])%', r'\1 %', clean_alert)
                    
                    # 根据警报类型和时间添加不同的样式
                    if '金叉' in clean_alert or '上穿' in clean_alert or '买入' in clean_alert or '看涨' in clean_alert:
                        html += f'<li class="golden-cross">{clean_alert}</li>'
                    elif '死叉' in clean_alert or '下穿' in clean_alert or '卖出' in clean_alert or '看跌' in clean_alert:
                        html += f'<li class="death-cross">{clean_alert}</li>'
                    elif 'RSI 超买' in clean_alert or 'RSI超买' in clean_alert:
                        html += f'<li class="death-cross">{clean_alert}</li>'
                    elif 'RSI 超卖' in clean_alert or 'RSI超卖' in clean_alert:
                        html += f'<li class="golden-cross">{clean_alert}</li>'
                    else:
                        # 如果不是交叉信号，使用普通样式
                        html += f'<li>{clean_alert}</li>'
                html += """</ul>
                    </div>
                </div>"""
            
            html += "</div>\n</div>"
        
        # HTML尾部
        html += """
            <footer style="margin-top: 40px; padding-top: 20px; border-top: 1px solid #E4E7EC; text-align: center; color: #8A919E; font-size: 0.9rem;">
                <p>EMA 股票分析工具 © 版权所有</p>
            </footer>
        </body>
        </html>
        """
        
        # 保存HTML报告
        with open(os.path.join(output_dir, 'analysis_results.html'), 'w', encoding='utf-8') as f:
            f.write(html)
            
    except Exception as e:
        print(f"{Colors.RED}Error saving results to HTML: {str(e)}{Colors.END}")
