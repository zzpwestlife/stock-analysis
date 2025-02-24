"""报告生成模块"""

import os
import base64
import re
import matplotlib.pyplot as plt
from .constants import Colors, EMA_PERIODS
import pandas as pd
from openpyxl.utils import get_column_letter

def remove_ansi_colors(text):
    """移除文本中的 ANSI 颜色代码
    Args:
        text (str): 包含 ANSI 颜色代码的文本
    Returns:
        str: 移除了 ANSI 颜色代码的文本
    """
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    return ansi_escape.sub('', text)

def save_analysis_plot(data, symbol, output_dir):
    """保存分析图表
    Args:
        data (pd.DataFrame): 股票数据
        symbol (str): 股票代码
        output_dir (str): 输出目录
    """
    try:
        plt.figure(figsize=(15, 8))  # 增加图表大小
        
        # 创建主图和副图
        ax1 = plt.gca()
        ax2 = ax1.twinx()
        
        # 格式化日期
        data_plot = data.copy()
        data_plot.index = pd.to_datetime(data_plot.index).strftime('%Y-%m-%d')
        
        # 绘制价格和均线
        ax1.plot(data_plot.index, data_plot['Close'], label='Price', color='black', linewidth=2, alpha=0.7)
        colors = ['blue', 'orange', 'red']  # Colors for 5, 50, 200 day EMAs
        for i, period in enumerate(EMA_PERIODS):
            column_name = f'EMA_{period}'
            ax1.plot(data_plot.index, data_plot[column_name], 
                    label=f'EMA {period}',
                    color=colors[i],
                    alpha=0.7)
        
        # Plot RSI
        ax2.plot(data_plot.index, data_plot['RSI'], label='RSI', color='gray', linewidth=1.5, alpha=0.7)
        ax2.axhline(y=70, color='r', linestyle='--', alpha=0.5)
        ax2.axhline(y=30, color='g', linestyle='--', alpha=0.5)
        
        # Set title and customize plot
        latest_date = pd.to_datetime(data_plot.index[-1]).strftime("%Y-%m-%d")
        ax1.set_title(f'Technical Analysis - {latest_date}', pad=20)
        ax1.grid(True, alpha=0.3)
        ax2.grid(True, alpha=0.3)
        
        # Format x-axis
        ax2.xaxis.set_major_locator(plt.MaxNLocator(10))
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')
        
        # Set labels
        ax1.set_ylabel('Price')
        ax2.set_ylabel('RSI')
        
        # Add legends
        ax1.legend(loc='upper left')
        ax2.legend(loc='upper left')
        
        # 保存图表
        plt.savefig(os.path.join(output_dir, f'{symbol}_analysis_plot.png'), 
                   bbox_inches='tight', dpi=100)
        plt.close()
        
    except Exception as e:
        print(f"{Colors.RED}Error saving analysis plot for {symbol}: {str(e)}{Colors.END}")

def format_value(value):
    """格式化数值
    Args:
        value: 要格式化的值
    Returns:
        str: 格式化后的字符串
    """
    if isinstance(value, float):
        return f"{value:.2f}"
    if isinstance(value, pd.Series):
        return f"{value.iloc[0]:.2f}"
    return str(value)

def save_to_excel(results, output_dir):
    """保存数据到Excel文件
    Args:
        results (list): 分析结果列表
        output_dir (str): 输出目录
    """
    try:
        # 创建一个Excel写入器
        excel_file = os.path.join(output_dir, 'stock_analysis.xlsx')
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 创建概览表
            overview_data = []
            for result in results:
                if not result.get('error'):
                    overview_data.append({
                        '股票代码': result['symbol'],
                        '当前价格': format_value(result['price']),
                        '价格变化': format_value(result['price_change']),
                        '价格变化率(%)': format_value(result['price_change_pct']),
                        'RSI': format_value(result['rsi']),
                        '警报数量': len(result.get('alert_details', []))
                    })
            
            # 保存概览表
            if overview_data:
                overview_df = pd.DataFrame(overview_data)
                overview_df.set_index('股票代码', inplace=True)
                
                # 格式化数字列为两位小数
                float_columns = ['当前价格', '价格变化', '价格变化率(%)', 'RSI']
                for col in float_columns:
                    overview_df[col] = pd.to_numeric(overview_df[col], errors='ignore')
                
                overview_df.to_excel(writer, sheet_name='概览', float_format='%.2f')
                
                # 获取工作表
                worksheet = writer.sheets['概览']
                
                # 定义样式
                from openpyxl.styles import PatternFill, Font
                red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
                
                # 为每个单元格添加条件格式化
                for row in range(2, worksheet.max_row + 1):  # 从第2行开始（跳过标题）
                    # 价格变化和变化率
                    for col in ['C', 'D']:  # B是股票代码，C是价格变化，D是变化率
                        cell = worksheet[f'{col}{row}']
                        if cell.value:
                            value = float(cell.value)
                            if value > 0:
                                cell.fill = green_fill
                            elif value < 0:
                                cell.fill = red_fill
                    
                    # RSI
                    rsi_cell = worksheet[f'E{row}']
                    if rsi_cell.value:
                        rsi = float(rsi_cell.value)
                        if rsi >= 70:
                            rsi_cell.fill = red_fill
                        elif rsi <= 30:
                            rsi_cell.fill = green_fill
                    
                    # 警报数量
                    alert_cell = worksheet[f'F{row}']
                    if alert_cell.value and int(alert_cell.value) > 0:
                        alert_cell.fill = yellow_fill
            else:
                # 如果没有概览数据，创建一个空的概览表
                pd.DataFrame().to_excel(writer, sheet_name='概览')
            
            # 为每个股票创建详细数据表
            for result in results:
                if not result.get('error') and 'data' in result:
                    symbol = result['symbol']
                    data = result['data'].copy()
                    
                    # 格式化日期索引
                    data.index = data.index.strftime('%Y-%m-%d')
                    
                    # 重命名列
                    rename_map = {
                        'Open': '开盘价',
                        'High': '最高价',
                        'Low': '最低价',
                        'Close': '收盘价',
                        'Adj Close': '调整后收盘价',
                        'Volume': '成交量',
                        'RSI': 'RSI'
                    }
                    for period in EMA_PERIODS:
                        rename_map[f'EMA_{period}'] = f'{period}日均线'
                    
                    data.rename(columns=rename_map, inplace=True)
                    
                    # 保存到Excel，所有数字列保留两位小数
                    data.to_excel(writer, sheet_name=symbol, float_format='%.2f')
                    
                    # 获取工作表
                    worksheet = writer.sheets[symbol]
                    
                    # 为每个单元格添加条件格式化
                    for row in range(2, worksheet.max_row + 1):
                        # RSI
                        rsi_col = None
                        for idx, col in enumerate(worksheet[1], 1):  # 第一行是标题
                            if col.value == 'RSI':
                                rsi_col = idx
                                break
                        
                        if rsi_col:
                            rsi_cell = worksheet.cell(row=row, column=rsi_col)
                            if rsi_cell.value:
                                rsi = float(rsi_cell.value)
                                if rsi >= 70:
                                    rsi_cell.fill = red_fill
                                elif rsi <= 30:
                                    rsi_cell.fill = green_fill
            
            # 设置列宽
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for idx, col in enumerate(worksheet.columns, 1):
                    worksheet.column_dimensions[get_column_letter(idx)].width = 15
                    
    except Exception as e:
        print(f"{Colors.RED}Error saving Excel file: {str(e)}{Colors.END}")

def generate_html_report(results, output_dir):
    """生成HTML报告
    Args:
        results (list): 分析结果列表
        output_dir (str): 输出目录
    """
    try:
        # 保存Excel文件
        save_to_excel(results, output_dir)
        
        # HTML头部
        html = """
        <html>
        <head>
            <title>股票分析报告</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
                .up { color: #4CAF50; }
                .down { color: #f44336; }
                .alert { color: #ff9800; }
                .alert-text { color: #ff9800; font-weight: bold; }
                .golden-cross { color: #4CAF50; font-weight: bold; }
                .death-cross { color: #f44336; font-weight: bold; }
                img { max-width: 100%; height: auto; margin: 10px 0; }
            </style>
        </head>
        <body>
            <h1>股票分析报告</h1>
            <h2>概览</h2>
            <p>详细数据已保存到 <a href="stock_analysis.xlsx">stock_analysis.xlsx</a></p>
            <table>
                <tr>
                    <th>股票代码</th>
                    <th>当前价格</th>
                    <th>价格变化</th>
                    <th>RSI</th>
                    <th>警报数量</th>
                </tr>
        """
        
        # 添加概览表格数据
        for result in results:
            if result.get('error'):
                html += f"""
                <tr>
                    <td>{result['symbol']}</td>
                    <td colspan="4" style="color: red;">Error: {result['error']}</td>
                </tr>
                """
                continue
                
            price_change = float(format_value(result['price_change']))
            price_change_class = 'up' if price_change > 0 else 'down'
            
            html += f"""
            <tr>
                <td>{result['symbol']}</td>
                <td>{format_value(result['price'])}</td>
                <td class="{price_change_class}">{format_value(result['price_change'])} ({format_value(result['price_change_pct'])}%)</td>
                <td>{format_value(result['rsi'])}</td>
                <td class="{'alert' if len(result.get('alert_details', [])) > 0 else ''}">{len(result.get('alert_details', []))}</td>
            </tr>
            """
        
        html += """
            </table>
            <h2>详细分析</h2>
        """
        
        # 添加每个股票的详细分析
        for result in results:
            if result.get('error'):
                continue
                
            html += f"""
            <div class="stock-detail">
                <h3>{result['symbol']} 详细分析</h3>
                <img src="{result['symbol']}_analysis_plot.png" alt="{result['symbol']} 分析图表">
            """
            
            # 添加警报信息
            if result.get('alert_details', []):
                html += "<h4>警报信息</h4><ul>"
                for alert in result['alert_details']:
                    # 提取日期并检查是否在最近10天内
                    import re
                    from datetime import datetime, timedelta
                    date_match = re.search(r'发生于：(\d{4}-\d{2}-\d{2})', alert)
                    is_recent = False
                    if date_match:
                        alert_date = datetime.strptime(date_match.group(1), '%Y-%m-%d')
                        current_date = datetime.now()
                        days_diff = (current_date - alert_date).days
                        is_recent = days_diff <= 10
                    
                    # 根据警报类型和时间添加不同的样式
                    if '金叉' in alert and is_recent:
                        html += f'<li class="golden-cross">{alert}</li>'
                    elif '死叉' in alert and is_recent:
                        html += f'<li class="death-cross">{alert}</li>'
                    else:
                        # 如果不是最近的交叉信号，或者是其他类型的警报，使用普通样式
                        html += f'<li>{alert}</li>'
                html += "</ul>"
            
            html += "</div>"
        
        # HTML尾部
        html += """
        </body>
        </html>
        """
        
        # 保存HTML报告
        with open(os.path.join(output_dir, 'analysis_results.html'), 'w', encoding='utf-8') as f:
            f.write(html)
            
    except Exception as e:
        print(f"{Colors.RED}Error saving results to HTML: {str(e)}{Colors.END}")
